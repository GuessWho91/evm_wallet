import os
import time
from datetime import date
from pathlib import Path
import openpyxl


class W3Logger:

    def __init__(self):

        script_location = Path(__file__).absolute().parent
        today = date.today()

        folder_path = f'{script_location}/logs'
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        folder_path_results = f'{script_location}/results'
        if not os.path.exists(folder_path_results):
            os.makedirs(folder_path_results)

        self.file_path = f'{folder_path}/{today}.txt'
        if not os.path.exists(self.file_path):
            with open(self.file_path, 'w', encoding='utf-8') as file:
                file.write('')

        self.file_path_results = f'{folder_path_results}/result.xlsx'
        if not os.path.exists(self.file_path_results):
            wb = openpyxl.Workbook()
            wb.active.cell(row=1, column=1).value = "wallets"
            wb.save(self.file_path_results)

        self.wb_results = openpyxl.load_workbook(self.file_path_results)

    def success(self, msg):
        with open(self.file_path, 'a+', encoding='utf-8') as file:
            msg = self.get_message(msg, 'SUCCESS')
            file.write(msg)
            print(msg)

    def error(self, msg):
        with open(self.file_path, 'a+', encoding='utf-8') as file:
            msg = self.get_message(msg, 'ERROR')
            file.write(msg)
            print(msg)

    def info(self, msg):
        with open(self.file_path, 'a+', encoding='utf-8') as file:
            msg = self.get_message(msg, 'INFO')
            file.write(msg)
            print(msg)

    def warning(self, msg):
        with open(self.file_path, 'a+', encoding='utf-8') as file:
            msg = self.get_message(msg, 'WARNING')
            file.write(msg)
            print(msg)

    def skip(self):
        with open(self.file_path, 'a+', encoding='utf-8') as file:
            file.write('\n')

    def result(self, wallet_address, column_name, msg):

        sheet = self.wb_results.active

        columns = []
        wallets = []

        for column in sheet.iter_cols():  # Iterate through columns
            column_value = column[0].value
            columns.append(column_value)
            if column_value == "wallets":
                for cell in column:
                    wallets.append(cell.value)

        try:
            row = wallets.index(wallet_address) + 1
        except ValueError:
            row = sheet.max_row + 1
            sheet.cell(row=row, column=1).value = wallet_address

        try:
            col = columns.index(column_name) + 1
        except ValueError:
            col = sheet.max_column + 1
            sheet.cell(row=1, column=col).value = column_name

        sheet.cell(row=row, column=col).value = msg
        self.wb_results.save(self.file_path_results)

    @staticmethod
    def get_message(msg, msg_type):
        return '{} | {} | {}\n'.format(time.strftime("%H:%M:%S"), msg_type.ljust(8), msg)
